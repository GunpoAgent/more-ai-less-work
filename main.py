import sys
import os
import pandas as pd
import folium
import json
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QFileDialog, 
                             QMessageBox, QProgressBar, QTextEdit, QSplitter,
                             QCheckBox, QScrollArea, QFrame)
from PyQt5.QtWebEngineWidgets import QWebEngineView
from PyQt5.QtCore import QUrl, QThread, pyqtSignal, Qt
from geocoding import process_excel_file, save_dataframe
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# API 키 (README.md에서 가져옴)
KAKAO_REST_API_KEY = "f4ca9d9fc9096a0f4f26a636e9d0ee29"

class GeocodingThread(QThread):
    """지오코딩을 백그라운드에서 처리하는 스레드"""
    
    progress_update = pyqtSignal(str)
    finished_signal = pyqtSignal(pd.DataFrame, list)
    error_signal = pyqtSignal(str)
    
    def __init__(self, file_path: str):
        super().__init__()
        self.file_path = file_path
        
    def run(self):
        try:
            self.progress_update.emit("엑셀 파일을 읽는 중...")
            df = process_excel_file(self.file_path, KAKAO_REST_API_KEY)
            
            self.progress_update.emit("지오코딩 완료! 결과를 저장하는 중...")
            save_dataframe(df)
            
            # 실패한 항목들 수집
            failed_items = []
            for idx, row in df.iterrows():
                if pd.isna(row.get('geocoding', '')) or row.get('geocoding', '') == '':
                    if not pd.isna(row.get('address', '')) and row.get('address', '') != '':
                        failed_items.append({
                            'row_index': idx + 1,
                            'type': row.get('type', ''),
                            'address': row.get('address', ''),
                            'date': row.get('date', '')
                        })
            
            self.finished_signal.emit(df, failed_items)
            
        except Exception as e:
            self.error_signal.emit(str(e))

class MapGeocodingApp(QMainWindow):
    """메인 애플리케이션 클래스"""
    
    def __init__(self):
        super().__init__()
        self.current_df = None
        self.failed_items = []
        self.metadata_file = "app_metadata.json"
        self.parquet_file = "geocoded_data.parquet"
        self.color_map = {}
        self.type_checkboxes = {}
        self.init_ui()
        self.load_last_session()
        
    def init_ui(self):
        """UI 초기화"""
        self.setWindowTitle("지오코딩 & 지도 시각화 애플리케이션")
        self.setGeometry(100, 100, 1400, 800)
        
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 메인 레이아웃
        main_layout = QHBoxLayout(central_widget)
        
        # 스플리터로 좌우 분할
        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter)
        
        # 좌측 지도 영역
        self.create_map_panel(splitter)
        
        # 우측 컨트롤 패널
        self.create_control_panel(splitter)
        
        # 스플리터 비율 설정
        splitter.setSizes([1000, 400])
        
    def create_control_panel(self, parent):
        """우측 컨트롤 패널 생성"""
        control_widget = QWidget()
        control_layout = QVBoxLayout(control_widget)
        
        # 마지막 갱신 정보
        self.update_date_label = QLabel("마지막 갱신 날짜: -")
        self.update_date_label.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px;")
        control_layout.addWidget(self.update_date_label)
        
        self.update_file_label = QLabel("저장된 파일: -")
        self.update_file_label.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px;")
        control_layout.addWidget(self.update_file_label)
        
        # 제목
        title_label = QLabel("지오코딩 & 지도 시각화")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px;")
        control_layout.addWidget(title_label)
        
        # 기본 엑셀 양식 다운로드 버튼
        self.download_button = QPushButton("기본 엑셀 양식 DOWNLOAD")
        self.download_button.clicked.connect(self.download_template)
        control_layout.addWidget(self.download_button)
        
        # 파일 선택 버튼
        self.file_button = QPushButton("UPLOAD EXCEL")
        self.file_button.clicked.connect(self.select_file)
        control_layout.addWidget(self.file_button)
        
        # 선택된 파일 경로 표시
        self.file_label = QLabel("selected file: 없음")
        self.file_label.setWordWrap(True)
        control_layout.addWidget(self.file_label)
        
        # 진행 상황 표시
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        control_layout.addWidget(self.progress_bar)
        
        # 상태 메시지
        self.status_label = QLabel("준비됨")
        control_layout.addWidget(self.status_label)
        
        # 결과 텍스트 영역
        self.result_text = QTextEdit()
        self.result_text.setMaximumHeight(200)
        self.result_text.setPlaceholderText("Geocoding Result:")
        control_layout.addWidget(self.result_text)
        
        # Check Failed List 버튼
        self.check_failed_button = QPushButton("CHECK FAILED LIST")
        self.check_failed_button.clicked.connect(self.export_failed_data)
        self.check_failed_button.setEnabled(False)
        control_layout.addWidget(self.check_failed_button)
        
        control_layout.addStretch()
        parent.addWidget(control_widget)
        
    def create_map_panel(self, parent):
        """좌측 지도 패널 생성"""
        # 지도 패널을 위한 컨테이너 위젯
        map_container = QWidget()
        map_layout = QVBoxLayout(map_container)
        map_layout.setContentsMargins(0, 0, 0, 0)
        
        # 웹뷰
        self.web_view = QWebEngineView()
        map_layout.addWidget(self.web_view)
        
        # 범례 오버레이 위젯 생성
        self.create_legend_overlay(map_container)
        
        parent.addWidget(map_container)
        
    def create_legend_overlay(self, parent):
        #지도 위에 Type (범례) 오버레이 생성
        """ 
        legend_overlay (최상위 컨테이너)
        └── overlay_layout (VBoxLayout)
         ├── legend_title ("Marker Type:" 라벨)
         └── legend_scroll (스크롤 되는 영역)
                └── legend_widget (스크롤되는 내용 )
                    └──  select_all_checkbox
                  
        """
        # 전체 범례 컨테이너
        self.legend_overlay = QWidget(parent)
        self.legend_overlay.setStyleSheet("""
            QWidget {
                background-color: rgba(255, 255, 255, 240);
                border: 2px solid #666666;
                border-radius: 8px;
                padding: 5px;
            }
        """) # 범례 컨테이너 스타일 설정 (배경 투명하게 설정, 테두리 설정, 둥근 모서리 설정, 패딩 설정)
        
        # 범례 레이아웃
        overlay_layout = QVBoxLayout(self.legend_overlay)  #위젯들을 새로로 배치 
        overlay_layout.setContentsMargins(10, 5, 10, 5) # 위젯들 주변 여백 설정 ( 왼쪽, 위, 오른쪽, 아래 )
        overlay_layout.setSpacing(3) # 위젯들 사이 간격 설정
        
        # 제목
        legend_title = QLabel("Marker Type:")
        legend_title.setStyleSheet("font-weight: bold; font-size: 14px; margin-bottom: 5px;") # 제목 스타일 설정
        overlay_layout.addWidget(legend_title) # 제목 위젯을 범례 레이아웃에 추가
        
        # 스크롤 영역
        self.legend_scroll = QScrollArea() # 스크롤 영역 위젯 생성
        self.legend_scroll.setWidgetResizable(True) # 내부  위젯 크기 자동으로 조정 (True: 조정 가능, False: 조정 불가능)
        self.legend_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded) # 필요할때만 세로 스크롤바 표시 
        self.legend_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff) # 가로 스크롤바 항상 숨김
        self.legend_scroll.setStyleSheet("""
            QScrollArea { 
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                width: 12px;
                background: rgba(0,0,0,0);
            }
            QScrollBar::handle:vertical {
                background: #888;
                border-radius: 6px;
            }
        """) # 스크롤 영역 테두리 제거, 배경 투명하게 설정 , 세로 스크롤바 생성 
         
        self.legend_widget = QWidget()
        self.legend_layout = QVBoxLayout(self.legend_widget)
        self.legend_layout.setContentsMargins(0, 0, 0, 0)
        self.legend_layout.setSpacing(2)
        
        
        # 모두 선택 체크박스
        self.select_all_checkbox = QCheckBox("모두 선택") #QcheckBox : 체크 / 언체크 가능한 상자 생성 위젯
        self.select_all_checkbox.setChecked(True) # 체크박스 초기 상태 설정 (True: 체크됨, False: 체크되지 않음)
        self.select_all_checkbox.clicked.connect(self.toggle_all_types) # clicked 이벤트 발생시 - connect - self.toggle_all_types 함수 호출
        self.select_all_checkbox.setStyleSheet("font-weight: bold; color: #333;") # 체크박스 스타일 설정 )
        self.legend_layout.addWidget(self.select_all_checkbox) # addWidget : 모두선택 체크박스 위젯을 범례 레이아웃에 추가, QvBoxLayout 이므로 가장 위쪽 (V = vercal )에 배치
        
        # 스크롤 가능 영역에 각 Type( 범례 )위젯 추가
        self.legend_scroll.setWidget(self.legend_widget) # 스크롤 영역에 범례 위젯 설정
        overlay_layout.addWidget(self.legend_scroll) # 스크롤 영역 위젯을 범례 레이아웃에 추가
        
        # 전체 범례 창 크기 설정
        self.legend_overlay.resize(200, 400)  #가로, 세로
        self.legend_overlay.hide()  # 초기에는 숨김 (엑셀 데이터 분석 후 표시를 의도함)
        
        # 창 크기 변경 시 범례 창 위치 조정
    def resizeEvent(self, event): #event : 이벤트 객체 ( 여기서는 창 크기 변경 이벤트 ) 보통 old_size , new_size 두개의 크기 정보를 포함함
        super().resizeEvent(event) # 부모 클래스의 resizeEvent 메서드 호출 ( 기본 동작 유지 )
        self.position_legend_overlay() # 범례 창 위치 조정 함수 호출 , 이후 범례 창 크기도 update
        
        #기본 지도 로드 함수 (경기도 군포시 청백리길6 , 군포 시청을 기준점으로)
    def load_default_map(self):
        tiles = "http://mt0.google.com/vt/lyrs=m&hl=ko&x={x}&y={y}&z={z}"
        attr = "Google" # 기본 지도가 아닌 google map 사용
        map_folium = folium.Map( # folium 라이브러리를 사용하여 지도 생성
            location=[37.36163002691529, 126.93520593427584], # 군포 시청 좌표
            zoom_start=11, # 초기 zoom 수준 ( 숫자 클수록 확대 )
            tiles=tiles, # 지도 타일 유형 ( Google Map )
            attr=attr # 지도 속성 ( Google Map )
        )
        # 위에 만들어진 지도를 html 파일로 저장
        html_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "map.html")
        map_folium.save(html_file)
        self.web_view.load(QUrl.fromLocalFile(html_file)) # 위에 만들어진 html 파일을 Gui 상의 webview에 로드 ( 지도 표시 )
        
        # 범례 숨기기 ( 초기에는 숨김 ) Q. 위에 숨기지 않았나요 ? A. 이전에 저장된 데이터가 있으면 범례 표시 되었음
        if hasattr(self, 'legend_overlay'):
            self.legend_overlay.hide()
            
    # 엑셀 파일 선택 함수
    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName( self, "엑셀 파일 선택", "", "Excel files (*.xlsx *.xls)") 
        #QFileDialog : 파일 선택 대화상자 생성 위젯, getOpenFileName : 파일 선택 대화상자 표시 함수
        # 파일 선택 대화상자 표시, 파일 선택 후 파일 경로 반환
        # arg 1 : self (파일 선택창의 부모 창, 현재 창을 기준으로 대화상자 표시)
        # arg 2 : 대화상자 제목 (엑셀 파일 선택)
        # arg 3 : "", 초기 디렉토리 (빈 문자열이면 현재 디렉토리 또는 최근 사용 디렉토리)
        # arg 4 : 파일 확장자 필터 (Excel 파일 확장자 필터)

        if file_path: #파일이 실제로 선택 되었는지 확인 , 빈 문자열은 False로 반환되므로 , 선택을 안한 경우엔 실행되지 않음
            try:
                # 파일 존재 여부 확인
                if not os.path.exists(file_path):
                    self.status_label.setText("파일을 찾을 수 없습니다.")
                    return
                # 파일 형식 검증
                if not file_path.lower().endswith(('.xlsx', '.xls')):
                    self.status_label.setText("올바른 엑셀 파일이 아닙니다.")
                    return
                
                # 올바른 파일이 선택되었을시 
                self.current_file_path = file_path # 선택된 파일 경로 저장
                self.file_label.setText(f"selected file: {os.path.basename(file_path)}") # 선택된 파일 경로 표시
                self.status_label.setText("파일이 선택되었습니다.")
                # 자동으로 지오코딩 실행
                self.start_geocoding()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 선택 중 오류가 발생했습니다:\n{e}")
                return # 오류 발생 시 함수 종료 ( 오류 발생 시 파일 선택 창 닫힘 )
            
    # 기본 엑셀 양식 다운로드 함수      
    def download_template(self):
        template_file = "integrated.xlsx"
        
        if not os.path.exists(template_file):
            QMessageBox.warning(self, f"파일 없음", "기본 양식 파일이 존재하지 않습니다.")
            return
            
        # 저장할 위치 선택
        save_path, _ = QFileDialog.getSaveFileName( #save_path, _: 저장 경로만 사용하고 필터 정보는 무시
            self, 
            "기본 엑셀 양식 저장", 
            "basic_form.xlsx",
            "Excel files (*.xlsx)"
        )
        
       
        if save_path:  # if save_path: 사용자가 실제로 저장 경로를 선택했는지 확인
            try:
                import shutil # shutil : 파일 복사 모듈
                shutil.copy2(template_file, save_path) # 기본 양식 파일을 선택된 경로에 복사 
                #shutil.copy(): 기본 복사, shutil.copyfile(): 파일 내용만 복사 , shutil.copy2(): 파일 내용 + 메타데이터 복사 
                QMessageBox.information(self, "다운로드 완료", f"기본 양식이 다음 위치에 저장되었습니다:\n{save_path}")
            except Exception as e:
                QMessageBox.critical(self, "다운로드 실패", f"파일 저장 중 오류가 발생했습니다:\n{e}")
 

    # 주소 -> 좌표 (위도,경도) 변환         
    def start_geocoding(self):
        if not hasattr(self, 'current_file_path'): 
            # hasattr(self, 'current_file_path'): 객체에 특정 속성이 존재확인, self.current_file_path가 설정되지 않은 상태에서 접근하면 AttributeError 발생 방지
            QMessageBox.warning(self, "파일 선택", "먼저 파일을 선택해주세요.")
            return # 파일이 선택되어 있지 않으면 함수 종료
            
        # UI 업데이트 ( 파일 선택 버튼 비활성화, 진행 바 표시, 상태 메시지 업데이트, 결과 텍스트 초기화 )
        self.file_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # 무한 진행바
        self.progress_bar.setStyleSheet("""
            QProgressBar {
            border: none;
            border-radius: 15px;
            background-color: rgba(0, 0, 0, 50);
            color: white;
            font-weight: bold;
            }
    
            QProgressBar::chunk {
            border-radius: 15px;
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #00ffff, stop:0.5 #ff00ff, stop:1 #ffff00);
            }
        """)

        self.status_label.setText("Processing...")
        self.result_text.clear()
        
        # 백그라운드 스레드에서 지오코딩 실행
        self.geocoding_thread = GeocodingThread(self.current_file_path)
        self.geocoding_thread.progress_update.connect(self.update_progress) # 진행 상황 업데이트 시 호출되는 함수 연결
        self.geocoding_thread.finished_signal.connect(self.geocoding_finished) # 지오코딩 완료 시 호출되는 함수 연결
        self.geocoding_thread.error_signal.connect(self.geocoding_error) # 지오코딩 오류 발생 시 호출되는 함수 연결
        self.geocoding_thread.start() # 지오코딩 작업 스레드 시작

        # 진행 상황 업데이트
    def update_progress(self, message): 
        self.status_label.setText(message) # 진행 상황 메시지 업데이트
        
        # GeoCoding 완료 처리 함수
    def geocoding_finished(self, df, failed_items):
        self.current_df = df # df : dataFrame
        
        # 변환에 실패한 항목들을 저장
        processed_failed_items = []
        for item in failed_items:
            processed_item = item.copy()
            if 'date' in processed_item and hasattr(processed_item['date'], 'strftime'): #datetime 객체를 JSON에 저장시키기 위해 문자열로 변환
                processed_item['date'] = processed_item['date'].strftime('%Y-%m-%d %H:%M:%S')
            processed_failed_items.append(processed_item)
        
        self.failed_items = processed_failed_items
        
        # 마지막 갱신 정보 업데이트
        current_time = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        self.update_date_label.setText(f"마지막 갱신 날짜: {current_time}")
        self.update_file_label.setText(f"현재 지도 데이터 : {os.path.basename(self.current_file_path)}")
        
        # UI 업데이트
        self.progress_bar.setVisible(False) # 진행 바 숨기기
        self.file_button.setEnabled(True) # 파일 선택 버튼 활성화
        
        # Check Failed List 버튼 활성화 (실패한 항목이 있는 경우)
        self.check_failed_button.setEnabled(len(failed_items) > 0)
        
        # 지도에 마커 표시
        self.status_label.setText("Updating Map...")
        self.create_map_with_markers() # 지도에 마커 표시하는 함수 호출
        
        # 결과 표시
        total_count = len(df) # 총 데이터 개수
        success_count = len(df[df['geocoding'].notna() & (df['geocoding'] != '')]) # 성공한 데이터 개수 df['geocoding'].notna() & (df['geocoding'] != ''): 지오코딩 성공 조건, notna(): null이 아닌 값, !=' ' : 빈 문자열 아닌값
        failed_count = len(failed_items) # 실패한 데이터 개수
        
        result_text = f"Result:\n"
        result_text += f"총 데이터: {total_count}개\n"
        result_text += f"성공: {success_count}개\n"
        result_text += f"실패: {failed_count}개\n\n"
        result_text += "※ 주소변환 실패 원인 : 오탈자, 누락, 다중 검색 주소 \n"
        result_text += "※ CHECK FAILED LIST - 주소 변환에 실패한 항목 확인 \n"
        self.result_text.setPlainText(result_text)

        # 메타데이터 저장
        self.save_session_metadata()
        
    def geocoding_error(self, error_message):
        """지오코딩 오류 처리"""
        self.progress_bar.setVisible(False)
        self.file_button.setEnabled(True)
        self.status_label.setText("지오코딩 중 오류가 발생했습니다.")
        
        QMessageBox.critical(self, "오류", f"지오코딩 중 오류가 발생했습니다:\n{error_message}")
    
    def export_failed_data(self):
        """실패한 데이터를 빨간색으로 처리하여 엑셀 파일로 내보내기"""
        if not self.failed_items:
            QMessageBox.information(self, "정보", "실패한 지오코딩 항목이 없습니다.")
            return
            
        try:
            # 원본 파일 이름에서 확장자 제거
            original_filename = os.path.splitext(os.path.basename(self.current_file_path))[0]
            failed_filename = f"FAILED_{original_filename}.xlsx"
            
            # 저장할 위치 선택
            save_path, _ = QFileDialog.getSaveFileName(
                self, 
                "실패한 항목 저장", 
                failed_filename,
                "Excel files (*.xlsx)"
            )
            
            if save_path:
                # 원본 파일을 복사하여 새 파일 생성
                import shutil
                shutil.copy2(self.current_file_path, save_path)
                
                # 실패한 행들을 빨간색으로 처리
                wb = load_workbook(save_path)
                ws = wb.active
                
                # 빨간색 배경 스타일
                red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                
                # 실패한 행들에 빨간색 배경 적용
                for item in self.failed_items:
                    row_num = item['row_index'] + 1  # 헤더 고려
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_num, column=col).fill = red_fill
                
                wb.save(save_path)
                
                QMessageBox.information(
                    self, 
                    "내보내기 완료", 
                    f"실패한 {len(self.failed_items)}개 항목이 빨간색으로 표시되어 저장되었습니다:\n{save_path}"
                )
                
        except Exception as e:
            QMessageBox.critical(self, "내보내기 실패", f"파일 저장 중 오류가 발생했습니다:\n{e}")
    
    def update_legend_checkboxes(self):
        """범례 체크박스 업데이트"""
        # 기존 체크박스들 제거 (모두 선택과 구분선 제외)
        for i in reversed(range(self.legend_layout.count())):
            if i > 1:  # 모두 선택과 구분선 이후의 항목들만 제거
                child = self.legend_layout.takeAt(i)
                if child.widget():
                    child.widget().deleteLater()
        
        self.type_checkboxes.clear()
        
        if self.current_df is None:
            self.legend_overlay.hide()
            return
            
        # 색상 매핑 업데이트
        unique_types = self.current_df['type'].dropna().unique()
        available_colors = ['red', 'blue', 'green', 'purple', 'orange', 'darkred', 'lightred', 
                           'beige', 'darkblue', 'darkgreen', 'cadetblue', 'darkpurple', 
                           'white', 'pink', 'lightblue', 'lightgreen', 'gray', 'black', 'lightgray']
        
        self.color_map = {}
        for i, place_type in enumerate(unique_types):
            self.color_map[place_type] = available_colors[i % len(available_colors)]
        
        # 타입별 개수 계산
        type_counts = self.current_df[self.current_df['geocoding'].notna() & (self.current_df['geocoding'] != '')]['type'].value_counts()
        
        # 각 타입별 체크박스 생성
        for place_type, color in self.color_map.items():
            count = type_counts.get(place_type, 0)
            checkbox = QCheckBox(f"{place_type} ({count}개)")
            checkbox.setChecked(True)
            checkbox.clicked.connect(self.update_map_markers)
            checkbox.setStyleSheet("color: #333; font-size: 12px;")
            self.type_checkboxes[place_type] = checkbox
            self.legend_layout.addWidget(checkbox)
        
        # 범례 오버레이 표시 및 위치 조정
        self.legend_overlay.show()
        self.position_legend_overlay()
            
    def position_legend_overlay(self):
        """범례 오버레이 위치 조정"""
        if hasattr(self, 'legend_overlay') and hasattr(self, 'web_view'):
            # 지도 영역의 좌표 계산
            web_view_rect = self.web_view.geometry()
            parent_rect = self.legend_overlay.parent().geometry()
            
            # 왼쪽 아래 위치 (여백 20px)
            x = 20
            y = web_view_rect.height() - self.legend_overlay.height() - 20
            self.legend_overlay.move(x, y)
            
    def toggle_all_types(self):
        """모든 타입 체크박스 토글"""
        is_checked = self.select_all_checkbox.isChecked()
        for checkbox in self.type_checkboxes.values():
            checkbox.setChecked(is_checked)
        self.update_map_markers()
        
    def update_map_markers(self):
        """체크박스 상태에 따라 지도 마커 업데이트"""
        if self.current_df is None:
            return
            
        # 선택된 타입들 수집
        selected_types = []
        for place_type, checkbox in self.type_checkboxes.items():
            if checkbox.isChecked():
                selected_types.append(place_type)
        
        # 모두 선택 체크박스 상태 업데이트
        if len(selected_types) == len(self.type_checkboxes):
            self.select_all_checkbox.setChecked(True)
        elif len(selected_types) == 0:
            self.select_all_checkbox.setChecked(False)
        else:
            self.select_all_checkbox.setChecked(False)
        
        # 지도 업데이트
        self.create_filtered_map(selected_types)
        
    def create_filtered_map(self, selected_types):

        """선택된 타입들만 표시하는 지도 생성"""
        # 기본 지도 생성 (경기도 군포시청 중심)
        tiles = "http://mt0.google.com/vt/lyrs=m&hl=ko&x={x}&y={y}&z={z}"
        attr = "Google"
        map_folium = folium.Map(
            location=[37.36163002691529, 126.93520593427584], 
            zoom_start=11,
            tiles=tiles,
            attr=attr
        )
        
        # 마커 추가 (선택된 타입만)
        marker_count = 0
        for idx, row in self.current_df.iterrows():
            geocoding_data = row.get('geocoding', '')
            place_type = row.get('type', '기타')
            
            # 선택된 타입만 표시
            if place_type not in selected_types:
                continue
                
            if pd.notna(geocoding_data) and geocoding_data != '':
                try:
                    lat, lng = map(float, geocoding_data.split(','))
                    address = row.get('address', '')
                    date = row.get('date', '')
                    note = row.get('비고', '')
                    
                    # 색상 선택
                    color = self.color_map.get(place_type, 'gray')
                    
                    # 팝업 내용
                    popup_content = f"""
                    <b>타입:</b> {place_type}<br>
                    <b>주소:</b> {address}<br>
                    <b>날짜:</b> {date}<br>
                    <b>비고:</b> {note}
                    """
                    
                    folium.Marker(
                        [lat, lng],
                        popup=folium.Popup(popup_content, max_width=300),
                        tooltip=f"{place_type} {address}",
                        icon=folium.Icon(color=color, icon='info-sign')
                    ).add_to(map_folium)
                    
                    marker_count += 1
                    
                except Exception as e:
                    print(f"마커 생성 실패 - 행 {idx}: {e}")

        # 지도 저장 및 표시
        html_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "map.html")
        map_folium.save(html_file)
        self.web_view.load(QUrl.fromLocalFile(html_file))
        
    def create_map_with_markers(self):
        """지오코딩된 데이터로 지도에 마커 표시"""
        if self.current_df is None:
            QMessageBox.warning(self, "데이터 없음", "먼저 지오코딩을 실행해주세요.")
            return
        
        # 범례 체크박스 업데이트
        self.update_legend_checkboxes()
        
        # 모든 타입이 선택된 상태로 지도 생성
        selected_types = list(self.color_map.keys())
        self.create_filtered_map(selected_types)

    def save_session_metadata(self):
        """현재 세션 메타데이터 저장"""
        if hasattr(self, 'current_file_path'):
            metadata = {
                "last_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source_file": os.path.basename(self.current_file_path),
                "full_path": self.current_file_path,
                "parquet_file": self.parquet_file,
                "total_records": len(self.current_df) if self.current_df is not None else 0,
                "success_count": len(self.current_df[self.current_df['geocoding'].notna() & (self.current_df['geocoding'] != '')]) if self.current_df is not None else 0,
                "failed_items": self.failed_items
            }
            
            try:
                with open(self.metadata_file, 'w', encoding='utf-8') as f:
                    json.dump(metadata, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"메타데이터 저장 실패: {e}")
    
    def load_last_session(self):
        """마지막 세션 로드"""
        try:
            # 메타데이터 파일과 parquet 파일이 모두 존재하는지 확인
            if os.path.exists(self.metadata_file) and os.path.exists(self.parquet_file):
                
                # 메타데이터 로드
                with open(self.metadata_file, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                
                # parquet 파일 로드
                self.current_df = pd.read_parquet(self.parquet_file)
                
                # 실패한 항목들 로드
                self.failed_items = metadata.get('failed_items', [])
                
                # UI 업데이트
                self.update_date_label.setText(f"마지막 갱신 날짜: {metadata['last_update']}")
                self.update_file_label.setText(f"저장된 파일: {metadata['source_file']}")
                
                # Check Failed List 버튼 활성화 (실패한 항목이 있는 경우)
                self.check_failed_button.setEnabled(len(self.failed_items) > 0)
                
                # 결과 텍스트 업데이트
                result_text = f"Result:\n"
                result_text += f"=== 저장된 데이터 로드 ===\n"
                result_text += f"Total Data: {metadata['total_records']}\n"
                result_text += f"Success: {metadata['success_count']}\n"
                result_text += f"Failed: {metadata['total_records'] - metadata['success_count']}\n\n"
                
                if self.failed_items:
                    result_text += f"=== 실패한 지오코딩 항목 ===\n"
                    for item in self.failed_items:
                        result_text += f"Row {item['row_index']}: [{item['type']}] {item['address']}\n"
                    result_text += f"\n"
                
                result_text += f"Data loaded from {self.parquet_file} file."
                
                self.result_text.setPlainText(result_text)
                
                # 자동으로 지도에 마커 표시
                self.create_map_with_markers()
                
                total_records = metadata['total_records']
                success_count = metadata['success_count']
                failed_count = total_records - success_count

                self.status_label.setText(f"표시된 마커 : {success_count}\n 총 데이터 : {total_records}\n 실패 : {failed_count}")
                
                return True
                
        except Exception as e:
            print(f"세션 로드 실패: {e}")
        
        # 세션 로드에 실패하면 기본 지도 로드
        self.load_default_map()
        return False

def main():
    app = QApplication(sys.argv)
    
    # 애플리케이션 정보 설정
    app.setApplicationName("지오코딩 & 지도 시각화")
    app.setApplicationVersion("1.0")
    
    window = MapGeocodingApp()
    window.show()
    
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
